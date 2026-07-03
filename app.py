import streamlit as st
import pandas as pd
import io
import urllib.request
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide", page_title="마케팅본부 월간 보고서")

st.markdown("""
<style>
* { font-family: 'Malgun Gothic', 'Apple SD Gothic Neo', sans-serif; }
.report-header {
    background: linear-gradient(135deg, #1a3a6b 0%, #2d5fa8 100%);
    color: white; padding: 16px 28px; font-size: 22px; font-weight: 900;
    margin-bottom: 20px; border-radius: 6px;
}
.box-title {
    display: inline-block; border: 2px solid #c0392b; border-radius: 20px;
    padding: 4px 18px; font-size: 15px; font-weight: 800;
    color: #c0392b; margin: 20px 0 10px 0;
}
.note-right { font-size: 12px; color: #555; text-align: right; margin-bottom: 4px; }
table { border-collapse: collapse; width: 100%; margin-bottom: 12px; }
th {
    background: #1e3a6b; color: white; font-weight: 700;
    border: 1px solid #2d5fa8; padding: 7px 10px;
    text-align: center; font-size: 13px; vertical-align: middle;
}
th.th-sub { background: #2d5fa8; }
td {
    border: 1px solid #ddd; padding: 7px 10px;
    text-align: center; font-size: 13px;
    vertical-align: middle; background: #ffffff;
}
tr:nth-child(even) td { background: #f8f9fc; }
td.lbl  { background: #dce6f5 !important; font-weight: 700; color: #1e3a6b; text-align: center !important; vertical-align: middle; }
td.slbl { background: #eef2fa !important; font-weight: 600; color: #333; text-align: center !important; vertical-align: middle; }
[data-testid="stSidebar"] { background: #f0f4fa; }
</style>
""", unsafe_allow_html=True)

# ── 헬퍼 ──────────────────────────────────────────────
NUM  = "style='color:#222;font-size:13px;font-weight:normal'"
SNUM = "style='color:#444;font-size:11px;font-weight:normal'"

def fmt(v, dec=0):
    if v is None or (isinstance(v, float) and pd.isna(v)): return "-"
    try:
        f = float(v)
        return f"{f:,.{dec}f}" if dec > 0 else f"{int(round(f)):,}"
    except: return str(v)

def fv(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    try: return float(v)
    except: return None

def n(v, dec=0):
    return f"<span {NUM}>{fmt(v,dec)}</span>"

def rate_html(a, p):
    try:
        val = float(a or 0) / float(p or 1) * 100
        return f"<span style='color:#222;font-size:13px;font-weight:normal'>{val:.1f}%</span>"
    except: return "<span style='color:#222;font-size:13px'>-</span>"

def rate_val(a, p):
    try: return float(a or 0) / float(p or 1) * 100
    except: return None

def safe(df, r, c):
    try:
        v = df.iloc[r, c]
        return None if (isinstance(v, float) and pd.isna(v)) else v
    except: return None

def d_inc(a, p):
    try: return fmt(float(a or 0) - float(p or 0))
    except: return "-"

def d_inc_v(a, p):
    try: return float(a or 0) - float(p or 0)
    except: return None

def cell(당월, 누계):
    return f"<span {NUM}>{fmt(당월)}</span><br><span {SNUM}>({fmt(누계)})</span>"

def rate_cell(da, dp, ca, cp):
    return f"{rate_html(da,dp)}<br><span {SNUM}>({rate_html(ca,cp)})</span>"

def inc_cell(da, dp, ca, cp):
    return f"<span {NUM}>{d_inc(da,dp)}</span><br><span {SNUM}>({d_inc(ca,cp)})</span>"

입력필요 = "<span style='color:#aaa;font-size:11px'>입력필요</span>"

# ── GitHub 설정 ──────────────────────────────────────
BASE  = "https://raw.githubusercontent.com/Han11112222/New-Volume-Monthly-Tracker/main"
GSHEET_URL = ("https://docs.google.com/spreadsheets/d/"
              "13HrIz6OytYDykXeXzXJ02I6XbaKin1YaKBoO2kBd6Bs/export?format=csv&gid=0")

# ── GitHub 파일 로드 ──────────────────────────────────
@st.cache_data(ttl=3600)
def load_github_plans():
    """new_1(계획), new_2(계획/실적) GitHub에서 로드"""
    errors, result = [], {}
    files = {
        "new_1": {
            "fname": "new_1.(작성용)6월 영업현황 보고(20260626).xlsx",
            "sheets": ["3_1. 개발량 계획", "(회의자료 입력용)공급전 및 공급량 현황"]
        },
        "new_2": {
            "fname": "new_2.(통합)신규개발량(202606)_5월 확정공급량적용_20260626.xlsx",
            "sheets": ["3_1. 개발량 계획", "3_2. 개발량 실적"]
        }
    }
    for key, info in files.items():
        try:
            url = f"{BASE}/{urllib.parse.quote(info['fname'])}"
            buf = io.BytesIO(urllib.request.urlopen(url, timeout=15).read())
            result[key] = {}
            for sh in info["sheets"]:
                try:
                    result[key][sh] = pd.read_excel(buf, sheet_name=sh, header=None)
                    buf.seek(0)
                except: pass
        except Exception as e:
            errors.append(f"{key}: {e}")
    return result, errors

@st.cache_data(ttl=3600)
def load_github_file(fname):
    """GitHub에서 특정 파일 로드"""
    try:
        url = f"{BASE}/{urllib.parse.quote(fname)}"
        return io.BytesIO(urllib.request.urlopen(url, timeout=15).read()), None
    except Exception as e:
        return None, str(e)

# ── 구글시트 ──────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def load_gsheet():
    try:
        req = urllib.request.Request(GSHEET_URL, headers={"User-Agent":"Mozilla/5.0"})
        raw = urllib.request.urlopen(req, timeout=15).read()
        df  = pd.read_csv(io.BytesIO(raw), header=0).iloc[:,:6]
        df.columns = ['일자','공급량_MJ','공급량_M3','평균기온','최저기온','최고기온']
        df['일자'] = pd.to_datetime(df['일자'], errors='coerce')
        df = df.dropna(subset=['일자'])
        df['공급량_MJ'] = pd.to_numeric(df['공급량_MJ'].astype(str).str.replace(',','').str.strip(), errors='coerce')
        df['공급량_M3'] = pd.to_numeric(df['공급량_M3'].astype(str).str.replace(',','').str.strip(), errors='coerce')
        df['GJ'] = df['공급량_MJ'] / 1000
        df['연'] = df['일자'].dt.year; df['월'] = df['일자'].dt.month
        monthly = df.groupby(['연','월']).agg(GJ=('GJ','sum'), MJ=('공급량_MJ','sum'), M3=('공급량_M3','sum')).reset_index()
        monthly['환산계수'] = monthly['MJ'] / monthly['M3']
        return monthly, None
    except Exception as e:
        return None, str(e)

def get_gj(mdf, yr, mo, cum=False):
    if mdf is None: return None
    if cum:
        f = mdf[(mdf['연']==yr) & (mdf['월']<=mo)]
        return float(f['GJ'].sum()) if len(f)>0 else None
    f = mdf[(mdf['연']==yr) & (mdf['월']==mo)]
    return float(f['GJ'].values[0]) if len(f)>0 else None

def get_환산계수(mdf, yr, mo):
    if mdf is None: return None
    f = mdf[(mdf['연']==yr) & (mdf['월']==mo)]
    return float(f['환산계수'].values[0]) if len(f)>0 else None

# ── 파일 파서 ─────────────────────────────────────────
@st.cache_data
def parse_xlsm(b, fn):
    buf = io.BytesIO(b); out = {}
    for sh in ['영업일보','공급전 계획(2026년)','공급량 계획_MJ(2026년)']:
        try: out[sh] = pd.read_excel(buf, sheet_name=sh, header=None, engine='openpyxl'); buf.seek(0)
        except: pass
    return out

@st.cache_data
def parse_dev(b, fn):
    buf = io.BytesIO(b)
    try:
        df = pd.read_excel(buf, sheet_name='Sheet1', header=0)
        df.columns = ['공급신청번호','시공업체','번지순번','주소','신청명','계약구분','건물구분',
                      '신청일','용도','업종','상품','등급','월사용예정량','공급승인일','공급일',
                      '서비스센터','설치장소주소','계량기번호','특정여부','월간개발량','공동주택명']
        return df
    except: return None

# ════════════════════════════════════════════════════
# 사이드바
# ════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 📋 보고서 설정")
    st.markdown("---")

    # 보고 연도/월 선택
    sel_year  = st.number_input("보고 연도", value=2026, step=1, format="%d")
    sel_month = st.selectbox("보고 월", list(range(1,13)), index=5)

    st.markdown("---")
    st.markdown("### 📁 파일 선택")
    st.markdown("**GitHub에서 자동 로드** 또는 직접 업로드")

    # 파일명 패턴
    yr = int(sel_year)
    mo = sel_month
    ym = f"{yr}{mo:02d}"  # 예: 202606

    # 영업일보 선택
    st.markdown(f"**① 영업일보**")
    xlsm_mode = st.radio("", ["GitHub 자동", "직접 업로드"], key="xlsm_mode", horizontal=True, label_visibility="collapsed")
    f_xlsm_upload = None
    xlsm_github_name = f"영업일보(월간)_{ym}.xlsm"
    if xlsm_mode == "직접 업로드":
        f_xlsm_upload = st.file_uploader("영업일보 파일", type=['xlsm','xlsx'], key="xlsm_up", label_visibility="collapsed")
    else:
        st.caption(f"📡 `{xlsm_github_name}`")

    st.markdown(f"**② 신규개발량**")
    dev_mode = st.radio("", ["GitHub 자동", "직접 업로드"], key="dev_mode", horizontal=True, label_visibility="collapsed")
    f_dev_upload = None
    dev_github_name = f"신규개발량_{ym}.xlsx"
    if dev_mode == "직접 업로드":
        f_dev_upload = st.file_uploader("신규개발량 파일", type=['xlsx'], key="dev_up", label_visibility="collapsed")
    else:
        st.caption(f"📡 `{dev_github_name}`")

    st.markdown("---")
    st.markdown("### 💡 총공급량 실적")
    st.caption("구글시트에서 자동 합산됩니다.")

# ════════════════════════════════════════════════════
# 데이터 로드
# ════════════════════════════════════════════════════
# 1. GitHub 계획 데이터
with st.spinner("📡 GitHub 계획 데이터 로드 중..."):
    gh, gh_err = load_github_plans()
    mdf, gs_err = load_gsheet()

auto_act = get_gj(mdf, yr, mo)
auto_cum = get_gj(mdf, yr, mo, cum=True)
환산계수  = get_환산계수(mdf, yr, mo)

# 2. 영업일보 로드
xlsm_bytes = None
xlsm_name  = ""
if xlsm_mode == "직접 업로드" and f_xlsm_upload:
    xlsm_bytes = f_xlsm_upload.read()
    xlsm_name  = f_xlsm_upload.name
else:
    buf, err = load_github_file(xlsm_github_name)
    if buf: xlsm_bytes = buf.read(); xlsm_name = xlsm_github_name

# 3. 신규개발량 로드
dev_bytes = None
dev_name  = ""
if dev_mode == "직접 업로드" and f_dev_upload:
    dev_bytes = f_dev_upload.read()
    dev_name  = f_dev_upload.name
else:
    buf, err = load_github_file(dev_github_name)
    if buf: dev_bytes = buf.read(); dev_name = dev_github_name

# ════════════════════════════════════════════════════
# 메인 화면
# ════════════════════════════════════════════════════
st.markdown(f'<div class="report-header">📊 마케팅본부 _ {yr}년 {mo}월 영업현황 보고서</div>',
            unsafe_allow_html=True)

# 로드 상태 표시 (작게)
col_s1, col_s2, col_s3 = st.columns(3)
with col_s1:
    if gh_err: st.warning("⚠️ GitHub 계획 로드 실패")
    else: st.success("✅ 계획 데이터 로드 완료")
with col_s2:
    if xlsm_bytes: st.success(f"✅ 영업일보: `{xlsm_name}`")
    else: st.warning(f"⚠️ 영업일보 없음: `{xlsm_github_name}`")
with col_s3:
    if dev_bytes: st.success(f"✅ 신규개발량: `{dev_name}`")
    else: st.warning(f"⚠️ 신규개발량 없음: `{dev_github_name}`")

if not xlsm_bytes:
    st.info(f"👈 사이드바에서 영업일보 파일을 업로드하거나, GitHub에 `{xlsm_github_name}` 파일을 올려주세요.")
    st.stop()

# ── 파싱 ─────────────────────────────────────────────
sheets = parse_xlsm(xlsm_bytes, xlsm_name)
dev_df = parse_dev(dev_bytes, dev_name) if dev_bytes else None

df_il   = sheets.get('영업일보')
df_vol  = sheets.get('공급량 계획_MJ(2026년)')
df_n1p  = gh.get("new_1",{}).get("3_1. 개발량 계획")
df_n1rt = gh.get("new_1",{}).get("(회의자료 입력용)공급전 및 공급량 현황")
df_n2p  = gh.get("new_2",{}).get("3_1. 개발량 계획")
df_n2r  = gh.get("new_2",{}).get("3_2. 개발량 실적")

m  = mo - 1
nc = m + 2
vc = m + 3

def s(df, r, c): return safe(df, r, c) if df is not None else None

# 영업일보 실적
RI = {'합계':7,'공동':8,'단독':9,'일반':10,'업무':11,'산업':12,'열병합':13}
def il(k, c): return s(df_il, RI.get(k), c)

공동_a=il('공동',4); 단독_a=il('단독',4); 소계_a=(공동_a or 0)+(단독_a or 0)
일반_a=il('일반',4); 업무_a=il('업무',4); 산업_a=il('산업',4)
열병_a=il('열병합',4); 합계_a=il('합계',4); 신규_당실=합계_a

공동_ca=il('공동',10); 단독_ca=il('단독',10); 소계_ca=(공동_ca or 0)+(단독_ca or 0)
일반_ca=il('일반',10); 업무_ca=il('업무',10); 산업_ca=il('산업',10)
열병_ca=il('열병합',10); 합계_ca=il('합계',10); 신규_누실=합계_ca

폐전_당실=il('합계',5); 폐전_누실=il('합계',11)
순증_당실=(float(신규_당실 or 0)-float(폐전_당실 or 0)) if 신규_당실 else None
순증_누실=(float(신규_누실 or 0)-float(폐전_누실 or 0)) if 신규_누실 else None

# 계획 (GitHub new_1)
신규_연간=s(df_n1p,15,14); 신규_당계=s(df_n1p,15,nc); 신규_누계=s(df_n1p,16,nc)
폐전_연간=s(df_n1p,34,14); 폐전_당계=s(df_n1p,34,nc); 폐전_누계=s(df_n1p,35,nc)
순증_연간=(신규_연간 or 0)-(폐전_연간 or 0)
순증_당계=(신규_당계 or 0)-(폐전_당계 or 0)
순증_누계=(신규_누계 or 0)-(폐전_누계 or 0)

개발량_연간=(s(df_n2p,104,14) or 0)/1000
개발량_당계=(s(df_n2p,103,nc) or 0)/1000
개발량_누계=(s(df_n2p,105,nc) or 0)/1000

# 신규개발량 실적: 2개 파일 + 구글시트 환산계수로 계산
if dev_df is not None and 환산계수:
    공동_m3 = float(공동_a or 0) * 48.5
    단독_m3 = float(단독_a or 0) * 43.4
    일반_m3 = float(일반_a or 0) * 203.3
    업무_m3 = float(dev_df[dev_df['용도']=='업무용']['월간개발량'].sum())
    산업_m3 = float(dev_df[dev_df['용도']=='산업용']['월간개발량'].sum())
    열병_m3 = float(dev_df[dev_df['용도'].str.contains('열병합', na=False)]['월간개발량'].sum())
    전체_m3 = 공동_m3+단독_m3+일반_m3+업무_m3+산업_m3+열병_m3
    개발량_당실 = 전체_m3 * 환산계수 / 1000
else:
    개발량_당실 = None

# 누계실적: new_2 row94(누적) col(nc) = 1월~당월 전체 누적
# row94 col(nc)에 당월까지의 누적값이 직접 들어있음
_nv = s(df_n2r, 94, nc)
개발량_누실 = float(_nv)/1000 if _nv else 개발량_당실

총공_연간=s(df_n1rt,10,2)
총공_당계=(s(df_vol,5,vc) or 0)/1000
총공_누계=(s(df_vol,6,vc) or 0)/1000
총공_당실_v = int(round(auto_act)) if auto_act else None
총공_누실_v = int(round(auto_cum)) if auto_cum else None

공동_p=s(df_n1p,2,nc); 단독_p=s(df_n1p,4,nc); 소계_p=(공동_p or 0)+(단독_p or 0)
일반_p=s(df_n1p,7,nc); 업무_p=s(df_n1p,9,nc); 산업_p=s(df_n1p,11,nc)
열병_p=s(df_n1p,13,nc); 합계_p=신규_당계
공동_cp=s(df_n1p,3,nc); 단독_cp=s(df_n1p,5,nc); 소계_cp=(공동_cp or 0)+(단독_cp or 0)
일반_cp=s(df_n1p,8,nc); 업무_cp=s(df_n1p,10,nc); 산업_cp=s(df_n1p,12,nc)
열병_cp=s(df_n1p,14,nc); 합계_cp=신규_누계

# ════════════════════════════════════════════════════
# 표1: 공급전 및 공급량 현황
# ════════════════════════════════════════════════════
st.markdown('<div class="box-title">📋 공급전 및 공급량 현황</div>', unsafe_allow_html=True)
st.markdown(f"""
<table>
  <thead>
    <tr>
      <th colspan="2" rowspan="2">구 분</th>
      <th rowspan="2">연간계획</th>
      <th colspan="3">{mo}월 (당월)</th>
      <th colspan="3">{mo}월 (누계)</th>
    </tr>
    <tr>
      <th class="th-sub">계획</th><th class="th-sub">실적</th><th class="th-sub">달성률</th>
      <th class="th-sub">계획</th><th class="th-sub">실적</th><th class="th-sub">달성률</th>
    </tr>
  </thead>
  <tbody>
    <tr>
      <td class="lbl" rowspan="3">공급전<br>(전)</td>
      <td class="slbl">신규개발전</td>
      <td>{n(신규_연간)}</td><td>{n(신규_당계)}</td><td>{n(신규_당실)}</td>
      <td>{rate_html(신규_당실,신규_당계)}</td>
      <td>{n(신규_누계)}</td><td>{n(신규_누실)}</td>
      <td>{rate_html(신규_누실,신규_누계)}</td>
    </tr>
    <tr>
      <td class="slbl">폐전</td>
      <td>{n(폐전_연간)}</td><td>{n(폐전_당계)}</td><td>{n(폐전_당실)}</td>
      <td>{rate_html(폐전_당실,폐전_당계)}</td>
      <td>{n(폐전_누계)}</td><td>{n(폐전_누실)}</td>
      <td>{rate_html(폐전_누실,폐전_누계)}</td>
    </tr>
    <tr>
      <td class="slbl">순증가</td>
      <td>{n(순증_연간)}</td><td>{n(순증_당계)}</td><td>{n(순증_당실)}</td>
      <td>{rate_html(순증_당실,순증_당계)}</td>
      <td>{n(순증_누계)}</td><td>{n(순증_누실)}</td>
      <td>{rate_html(순증_누실,순증_누계)}</td>
    </tr>
    <tr>
      <td class="lbl" rowspan="2">공급량<br>(GJ)</td>
      <td class="slbl">신규개발량</td>
      <td>{n(개발량_연간)}</td><td>{n(개발량_당계)}</td>
      <td>{n(개발량_당실) if 개발량_당실 else 입력필요}</td>
      <td>{rate_html(개발량_당실,개발량_당계) if 개발량_당실 else n(None)}</td>
      <td>{n(개발량_누계)}</td>
      <td>{n(개발량_누실) if 개발량_누실 else n(None)}</td>
      <td>{rate_html(개발량_누실,개발량_누계) if 개발량_누실 else n(None)}</td>
    </tr>
    <tr>
      <td class="slbl">총공급량</td>
      <td>{n(총공_연간)}</td><td>{n(총공_당계)}</td>
      <td>{n(총공_당실_v) if 총공_당실_v else 입력필요}</td>
      <td>{rate_html(총공_당실_v,총공_당계) if 총공_당실_v else n(None)}</td>
      <td>{n(총공_누계)}</td>
      <td>{n(총공_누실_v) if 총공_누실_v else 입력필요}</td>
      <td>{rate_html(총공_누실_v,총공_누계) if 총공_누실_v else n(None)}</td>
    </tr>
  </tbody>
</table>""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════
# 표2: 신규개발전 상세
# ════════════════════════════════════════════════════
st.markdown(f'<div class="box-title">📋 {mo}월 신규개발전 상세 현황</div>', unsafe_allow_html=True)
st.markdown('<div class="note-right">(단위 : 전)</div>', unsafe_allow_html=True)
st.markdown(f"""
<table>
  <thead>
    <tr>
      <th rowspan="2">구 분</th>
      <th colspan="3">주택용</th>
      <th rowspan="2">일반용</th><th rowspan="2">업무용</th>
      <th rowspan="2">산업용</th><th rowspan="2">열병합</th><th rowspan="2">합계</th>
    </tr>
    <tr>
      <th class="th-sub">공동주택</th><th class="th-sub">단독주택</th><th class="th-sub">소계</th>
    </tr>
  </thead>
  <tbody>
    <tr>
      <td class="lbl">계획</td>
      <td>{cell(공동_p,공동_cp)}</td><td>{cell(단독_p,단독_cp)}</td><td>{cell(소계_p,소계_cp)}</td>
      <td>{n(일반_p)}</td><td>{n(업무_p)}</td>
      <td>{n(산업_p)}</td><td>{n(열병_p)}</td><td>{cell(합계_p,합계_cp)}</td>
    </tr>
    <tr>
      <td class="lbl">실적</td>
      <td>{cell(공동_a,공동_ca)}</td><td>{cell(단독_a,단독_ca)}</td><td>{cell(소계_a,소계_ca)}</td>
      <td>{cell(일반_a,일반_ca)}</td><td>{cell(업무_a,업무_ca)}</td>
      <td>{cell(산업_a,산업_ca)}</td><td>{cell(열병_a,열병_ca)}</td><td>{cell(합계_a,합계_ca)}</td>
    </tr>
    <tr>
      <td class="lbl">달성률</td>
      <td>{rate_cell(공동_a,공동_p,공동_ca,공동_cp)}</td>
      <td>{rate_cell(단독_a,단독_p,단독_ca,단독_cp)}</td>
      <td>{rate_cell(소계_a,소계_p,소계_ca,소계_cp)}</td>
      <td>{rate_cell(일반_a,일반_p,일반_ca,일반_cp)}</td>
      <td>{rate_cell(업무_a,업무_p,업무_ca,업무_cp)}</td>
      <td>{rate_cell(산업_a,산업_p,산업_ca,산업_cp)}</td>
      <td>{rate_cell(열병_a,열병_p,열병_ca,열병_cp)}</td>
      <td>{rate_cell(합계_a,합계_p,합계_ca,합계_cp)}</td>
    </tr>
    <tr>
      <td class="lbl">증감</td>
      <td>{inc_cell(공동_a,공동_p,공동_ca,공동_cp)}</td>
      <td>{inc_cell(단독_a,단독_p,단독_ca,단독_cp)}</td>
      <td>{inc_cell(소계_a,소계_p,소계_ca,소계_cp)}</td>
      <td>{inc_cell(일반_a,일반_p,일반_ca,일반_cp)}</td>
      <td>{inc_cell(업무_a,업무_p,업무_ca,업무_cp)}</td>
      <td>{inc_cell(산업_a,산업_p,산업_ca,산업_cp)}</td>
      <td>{inc_cell(열병_a,열병_p,열병_ca,열병_cp)}</td>
      <td>{inc_cell(합계_a,합계_p,합계_ca,합계_cp)}</td>
    </tr>
  </tbody>
</table>
<p style="font-size:12px;color:#555;">※ (괄호)는 누계 기준임.</p>""", unsafe_allow_html=True)

# ── 산업용 신규 업체 ──────────────────────────────────
ind = biz = None
if dev_df is not None:
    ind = dev_df[dev_df['용도'].str.contains('산업', na=False)].copy()
    biz = dev_df[dev_df['용도'].str.contains('업무', na=False)].copy()

    def render_company_table(df_co, title):
        if df_co.empty:
            st.info(f"✅ {mo}월 {title} 없음"); return
        st.markdown(f'<div class="box-title">🏭 {mo}월 {title}</div>', unsafe_allow_html=True)
        rows_html = "".join([
            f"<tr><td>{r['신청명']}</td><td>{r['업종']}</td>"
            f"<td>{fmt(r['월사용예정량'])} ㎥</td>"
            f"<td>{fmt(r['월간개발량'],2)} GJ</td>"
            f"<td>{str(r['공급일'])[:10] if pd.notna(r['공급일']) else '-'}</td>"
            f"<td style='text-align:left;font-size:11px'>{r['주소']}</td></tr>"
            for _, r in df_co.iterrows()
        ])
        st.markdown(f"""
        <table>
          <thead><tr>
            <th>업체명</th><th>업종</th><th>월사용예정량(㎥)</th>
            <th>월사용예정량(GJ)</th><th>공급일</th><th>주소</th>
          </tr></thead>
          <tbody>{rows_html}
            <tr style="background:#dce6f5">
              <td colspan="2" style="font-weight:700">합 계</td>
              <td>{fmt(df_co['월사용예정량'].sum())} ㎥</td>
              <td>{fmt(df_co['월간개발량'].sum(),2)} GJ</td>
              <td colspan="2"></td>
            </tr>
          </tbody>
        </table>""", unsafe_allow_html=True)

    render_company_table(ind, "산업용 신규 업체 현황")
    render_company_table(biz, "업무용 신규 업체 현황")
else:
    st.warning("👈 사이드바에서 신규개발량 파일을 선택하면 산업용/업무용 업체 현황이 표시됩니다.")

# ── 엑셀 다운로드 ─────────────────────────────────────
st.markdown("---")

def make_excel(yr, mo, rows1, rows2, ind_df, biz_df):
    wb = Workbook(); ws = wb.active
    ws.title = f"{yr}년{mo}월 영업현황"
    hf = PatternFill("solid", fgColor="1E3A6B")
    sf = PatternFill("solid", fgColor="2D5FA8")
    lf = PatternFill("solid", fgColor="DCE6F5")
    nf = Font(name="맑은 고딕", size=10)
    hft= Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    lft= Font(name="맑은 고딕", bold=True, color="1E3A6B", size=10)
    ctr= Alignment(horizontal="center", vertical="center", wrap_text=True)
    th = Side(style="thin", color="CCCCCC")
    bd = Border(left=th,right=th,top=th,bottom=th)
    num_fmt = '#,##0'; pct_fmt = '0.0%'

    def hc(ws, r, c, val, fill=None, font=None, fmt=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = ctr; cell.border = bd
        cell.fill = fill if fill else PatternFill()
        cell.font = font if font else nf
        if fmt: cell.number_format = fmt
        return cell

    def pct_str(v):
        try: return f"{float(v):.1f}%" if v is not None else "-"
        except: return "-"
    def comma_str(v):
        try: return f"{int(round(float(v))):,}" if v is not None else "-"
        except: return "-"

    # 표1 헤더
    ws.merge_cells('A1:B2'); hc(ws,1,1,"구 분",hf,hft)
    ws.merge_cells('C1:C2'); hc(ws,1,3,"연간계획",hf,hft)
    ws.merge_cells('D1:F1'); hc(ws,1,4,f"{mo}월 (당월)",hf,hft)
    ws.merge_cells('G1:I1'); hc(ws,1,7,f"{mo}월 (누계)",hf,hft)
    for c,v in [(4,"계획"),(5,"실적"),(6,"달성률"),(7,"계획"),(8,"실적"),(9,"달성률")]:
        hc(ws,2,c,v,sf,hft)

    r = 3
    for 구분1, 구분2, 연간, 당계, 당실, 누계, 누실 in rows1:
        if 구분1:
            end = r+2 if 구분1=="공급전\n(전)" else r+1
            ws.merge_cells(start_row=r,start_column=1,end_row=end,end_column=1)
            hc(ws,r,1,구분1,lf,lft)
        hc(ws,r,2,구분2,PatternFill("solid",fgColor="EEF2FA"),Font(name="맑은 고딕",size=10,color="333333"))
        hc(ws,r,3,fv(연간),None,nf,num_fmt)
        hc(ws,r,4,fv(당계),None,nf,num_fmt)
        hc(ws,r,5,fv(당실),None,nf,num_fmt)
        rv = rate_val(당실,당계)
        hc(ws,r,6,rv/100 if rv else None,None,nf,pct_fmt)
        hc(ws,r,7,fv(누계),None,nf,num_fmt)
        hc(ws,r,8,fv(누실),None,nf,num_fmt)
        rv2 = rate_val(누실,누계)
        hc(ws,r,9,rv2/100 if rv2 else None,None,nf,pct_fmt)
        r += 1

    # 표2 헤더
    r += 1
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9)
    hc(ws,r,1,f"{mo}월 신규개발전 상세 현황 (단위: 전)",hf,hft); r+=1
    hc(ws,r,1,"구 분",hf,hft)
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
    hc(ws,r,2,"주택용",hf,hft)
    for c,v in [(5,"일반용"),(6,"업무용"),(7,"산업용"),(8,"열병합"),(9,"합계")]:
        hc(ws,r,c,v,hf,hft)
    r+=1
    for c,v in [(1,""),(2,"공동주택"),(3,"단독주택"),(4,"소계"),(5,""),(6,""),(7,""),(8,""),(9,"")]:
        hc(ws,r,c,v,sf,hft)
    r+=1

    def pct_str2(v): 
        try: return f"{float(v):.1f}%" if v is not None else "-"
        except: return "-"

    for 구분, *vals in rows2:
        pairs = list(zip(vals[::2], vals[1::2]))
        hc(ws,r,1,구분,lf,lft)
        for ci,(당월_v,누계_v) in enumerate(pairs):
            col = ci+2
            if 구분 == "달성률":
                val_str = pct_str2(당월_v)+"\n("+pct_str2(누계_v)+")"
            else:
                val_str = comma_str(당월_v)+"\n("+comma_str(누계_v)+")"
            c_obj = ws.cell(row=r,column=col,value=val_str)
            c_obj.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            c_obj.border = bd; c_obj.font = nf
        ws.row_dimensions[r].height = 30; r+=1

    def write_co(ws, r, title, df_co):
        if df_co is None or df_co.empty: return r
        r+=1
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        hc(ws,r,1,title,hf,hft); r+=1
        for c,v in enumerate(["업체명","업종","월사용예정량(㎥)","월사용예정량(GJ)","공급일","주소"],1):
            hc(ws,r,c,v,sf,hft)
        r+=1
        for _,row in df_co.iterrows():
            hc(ws,r,1,row['신청명'],None,nf); hc(ws,r,2,row['업종'],None,nf)
            hc(ws,r,3,fv(row['월사용예정량']),None,nf,num_fmt)
            hc(ws,r,4,fv(row['월간개발량']),None,nf,'#,##0.00')
            공일 = str(row['공급일'])[:10] if pd.notna(row['공급일']) else '-'
            hc(ws,r,5,공일,None,nf)
            c6=ws.cell(row=r,column=6,value=row['주소'])
            c6.alignment=Alignment(horizontal="left",vertical="center")
            c6.border=bd; c6.font=nf; r+=1
        hc(ws,r,1,"합 계",lf,lft); hc(ws,r,2,"",lf,lft)
        hc(ws,r,3,float(df_co['월사용예정량'].sum()),lf,lft,num_fmt)
        hc(ws,r,4,float(df_co['월간개발량'].sum()),lf,lft,'#,##0.00')
        return r+1

    r = write_co(ws,r,f"{mo}월 산업용 신규 업체 현황",ind_df)
    r = write_co(ws,r,f"{mo}월 업무용 신규 업체 현황",biz_df)

    for i,w in enumerate([10,14,12,12,12,12,12,12,12],1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

rows1 = [
    ("공급전\n(전)","신규개발전",신규_연간,신규_당계,신규_당실,신규_누계,신규_누실),
    (None,"폐전",폐전_연간,폐전_당계,폐전_당실,폐전_누계,폐전_누실),
    (None,"순증가",순증_연간,순증_당계,순증_당실,순증_누계,순증_누실),
    ("공급량\n(GJ)","신규개발량",개발량_연간,개발량_당계,개발량_당실,개발량_누계,개발량_누실),
    (None,"총공급량",총공_연간,총공_당계,총공_당실_v,총공_누계,총공_누실_v),
]
rows2 = [
    ("계획",공동_p,공동_cp,단독_p,단독_cp,소계_p,소계_cp,일반_p,일반_cp,업무_p,업무_cp,산업_p,산업_cp,열병_p,열병_cp,합계_p,합계_cp),
    ("실적",공동_a,공동_ca,단독_a,단독_ca,소계_a,소계_ca,일반_a,일반_ca,업무_a,업무_ca,산업_a,산업_ca,열병_a,열병_ca,합계_a,합계_ca),
    ("달성률",rate_val(공동_a,공동_p),rate_val(공동_ca,공동_cp),rate_val(단독_a,단독_p),rate_val(단독_ca,단독_cp),
              rate_val(소계_a,소계_p),rate_val(소계_ca,소계_cp),rate_val(일반_a,일반_p),rate_val(일반_ca,일반_cp),
              rate_val(업무_a,업무_p),rate_val(업무_ca,업무_cp),rate_val(산업_a,산업_p),rate_val(산업_ca,산업_cp),
              rate_val(열병_a,열병_p),rate_val(열병_ca,열병_cp),rate_val(합계_a,합계_p),rate_val(합계_ca,합계_cp)),
    ("증감",d_inc_v(공동_a,공동_p),d_inc_v(공동_ca,공동_cp),d_inc_v(단독_a,단독_p),d_inc_v(단독_ca,단독_cp),
            d_inc_v(소계_a,소계_p),d_inc_v(소계_ca,소계_cp),d_inc_v(일반_a,일반_p),d_inc_v(일반_ca,일반_cp),
            d_inc_v(업무_a,업무_p),d_inc_v(업무_ca,업무_cp),d_inc_v(산업_a,산업_p),d_inc_v(산업_ca,산업_cp),
            d_inc_v(열병_a,열병_p),d_inc_v(열병_ca,열병_cp),d_inc_v(합계_a,합계_p),d_inc_v(합계_ca,합계_cp)),
]

xl = make_excel(yr, mo, rows1, rows2, ind, biz)
col_dl = st.columns([2,1,2])
with col_dl[1]:
    st.download_button(
        label="📥 엑셀 파일로 저장",
        data=xl,
        file_name=f"{yr}년{mo}월_영업현황보고.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary"
    )
